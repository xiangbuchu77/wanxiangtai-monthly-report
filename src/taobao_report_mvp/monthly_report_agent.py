from __future__ import annotations

import argparse
import copy
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_TOTAL = Path("outputs/ecommerce-monthly/淘宝天猫运营标准经营总表_MVP.xlsx")
DEFAULT_TEMPLATE = Path("templates/万相台月报模板.xlsx")
DEFAULT_OUTPUT = Path("outputs/ecommerce-monthly/万相台AI分析月报.xlsx")
DEFAULT_PACKET = Path("outputs/ecommerce-monthly/ai_agent_packet.json")


AD_METRICS_CONTRACT = {
    "monthly_summary": ["月份", "广告引导成交金额", "总推广花费", "点击量", "总ROI"],
    "plan_detail": ["月份", "推广计划名称", "计划类型", "花费", "展现量", "点击量", "ROI"],
    "product_ad": ["月份", "商品ID", "商品名称", "广告花费", "广告ROI"],
}
CORE_METRIC_LABELS = [
    "店铺总访客数",
    "店铺总成交金额（元）",
    "店铺支付转化率",
    "广告总花费（元）",
    "广告带来的成交金额（元）",
    "广告投入产出比（ROI）",
    "平均点击成本（元）",
    "总点击量",
]


@dataclass(frozen=True)
class AgentInputs:
    total_path: Path
    template_path: Path
    ad_metrics_path: Path | None
    ai_json_path: Path | None
    output_path: Path
    packet_path: Path
    target_month: str | None


def main() -> None:
    parser = argparse.ArgumentParser(
        description="用标准经营总表和结构化广告数据填充万相台月报模板；AI只生成文字分析字段。"
    )
    parser.add_argument("--total", type=Path, default=DEFAULT_TOTAL)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--ad-metrics", type=Path, help="结构化广告导出表，需包含 monthly_summary/plan_detail/product_ad 数据")
    parser.add_argument("--ai-json", type=Path, help="AI返回的JSON文件；不传则只生成AI分析任务包")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--packet", type=Path, default=DEFAULT_PACKET)
    parser.add_argument("--month", help="目标月份，例如 2026-05；不传则使用经营总表最新月份")
    args = parser.parse_args()

    inputs = AgentInputs(
        total_path=args.total,
        template_path=args.template,
        ad_metrics_path=args.ad_metrics,
        ai_json_path=args.ai_json,
        output_path=args.output,
        packet_path=args.packet,
        target_month=args.month,
    )
    run_agent(inputs)


def run_agent(inputs: AgentInputs) -> None:
    sheets = pd.read_excel(inputs.total_path, sheet_name=None)
    month = inputs.target_month or str(sheets["经营总表"]["月份"].dropna().max())
    ad_metrics = load_ad_metrics(inputs.ad_metrics_path) if inputs.ad_metrics_path else {}
    validation = validate_required_data(sheets, ad_metrics, month)
    context = build_context(sheets, ad_metrics, month, validation)
    packet = build_ai_packet(context)
    inputs.packet_path.parent.mkdir(parents=True, exist_ok=True)
    inputs.packet_path.write_text(json.dumps(packet, ensure_ascii=False, indent=2), encoding="utf-8")

    if validation["blocking_missing"]:
        print(f"packet={inputs.packet_path}")
        print("status=blocked")
        print("missing=" + "、".join(validation["blocking_missing"]))
        return

    ai_text = load_ai_json(inputs.ai_json_path) if inputs.ai_json_path else empty_ai_text()
    fill_template(inputs.template_path, inputs.output_path, context, ai_text)
    print(f"packet={inputs.packet_path}")
    print(f"output={inputs.output_path}")
    print("status=ok")


def load_ad_metrics(path: Path) -> dict[str, pd.DataFrame]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        return {name: pd.DataFrame(rows) for name, rows in payload.items()}
    if suffix in {".xlsx", ".xls"}:
        sheets = pd.read_excel(path, sheet_name=None)
        return {normalize_sheet_name(name): frame for name, frame in sheets.items()}
    if suffix == ".csv":
        return {"monthly_summary": pd.read_csv(path)}
    raise ValueError(f"不支持的广告数据文件格式: {path}")


def normalize_sheet_name(name: str) -> str:
    mapping = {
        "月度汇总": "monthly_summary",
        "广告月度汇总": "monthly_summary",
        "计划明细": "plan_detail",
        "推广计划": "plan_detail",
        "商品广告": "product_ad",
        "商品推广": "product_ad",
    }
    return mapping.get(str(name).strip(), str(name).strip())


def validate_required_data(sheets: dict[str, pd.DataFrame], ad_metrics: dict[str, pd.DataFrame], month: str) -> dict[str, Any]:
    missing: list[str] = []
    warnings: list[str] = []
    core = month_scope(sheets.get("经营总表", pd.DataFrame()), month)
    product = month_scope(sheets.get("商品经营", pd.DataFrame()), month)

    for field in ["访客数", "支付金额", "支付转化率_计算"]:
        if core.empty or field not in core.columns or numeric_sum(core, field) <= 0:
            missing.append(f"缺少指标：{friendly_core_field(field)}")
    if core.empty or "推广花费合计" not in core.columns or numeric_sum(core, "推广花费合计") <= 0:
        warnings.append("缺少指标：广告总花费（元），广告预算模块会生成简版")
    if previous_month_row(sheets.get("经营总表", pd.DataFrame()), month).empty:
        warnings.append("缺少上期店铺收益数据，店铺收益环比会改为本期绝对表现分析")
    if product.empty:
        warnings.append("缺少本期商品经营明细，主要商品数据模块会省略或简化")

    monthly_ad = month_scope(ad_metrics.get("monthly_summary", pd.DataFrame()), month)
    previous_monthly_ad = previous_month_row(ad_metrics.get("monthly_summary", pd.DataFrame()), month)
    plan_ad = month_scope(ad_metrics.get("plan_detail", pd.DataFrame()), month)
    if not plan_ad.empty and "花费" in plan_ad.columns:
        plan_ad = plan_ad[pd.to_numeric(plan_ad["花费"], errors="coerce").fillna(0) > 0]
        plan_ad = plan_ad.sort_values("花费", ascending=False)
    product_ad = month_scope(ad_metrics.get("product_ad", pd.DataFrame()), month)
    if not product_ad.empty and "商品ID" in product_ad.columns:
        product_ad["商品ID"] = product_ad["商品ID"].map(normalize_id)

    for field in AD_METRICS_CONTRACT["monthly_summary"][1:]:
        if monthly_ad.empty or field not in monthly_ad.columns or numeric_sum(monthly_ad, field) <= 0:
            warnings.append(f"缺少指标：{friendly_ad_field(field)}，预算、ROI、CPC模块会生成简版")
    if previous_monthly_ad.empty:
        warnings.append("缺少上期广告指标，广告环比分析会弱化")

    for field in ["推广计划名称", "计划类型", "花费", "点击量", "ROI"]:
        if plan_ad.empty or field not in plan_ad.columns:
            warnings.append(f"缺少推广计划明细：{field}")
    if not plan_ad.empty and "点击量" in plan_ad.columns and numeric_sum(plan_ad, "点击量") <= 0:
        warnings.append("推广计划明细缺少有效点击量")
    if not plan_ad.empty and "花费" in plan_ad.columns and numeric_sum(plan_ad, "花费") <= 0:
        warnings.append("推广计划明细缺少有效花费")

    for field in ["商品ID", "商品名称", "广告花费", "广告ROI"]:
        if product_ad.empty or field not in product_ad.columns:
            warnings.append(f"缺少商品广告数据：{field}")

    return {
        "blocking_missing": sorted(set(missing)),
        "warnings": sorted(set(warnings)),
        "month": month,
    }


def friendly_core_field(field: str) -> str:
    mapping = {
        "访客数": "店铺总访客数",
        "支付金额": "店铺总成交金额（元）",
        "支付转化率_计算": "店铺支付转化率",
    }
    return mapping.get(field, field)


def friendly_ad_field(field: str) -> str:
    mapping = {
        "总推广花费": "广告总花费（元）",
        "广告引导成交金额": "广告带来的成交金额（元）",
        "总ROI": "广告投入产出比（ROI）",
        "点击量": "总点击量",
    }
    return mapping.get(field, field)


def build_context(
    sheets: dict[str, pd.DataFrame],
    ad_metrics: dict[str, pd.DataFrame],
    month: str,
    validation: dict[str, Any],
) -> dict[str, Any]:
    core_all = sheets["经营总表"].copy()
    core = month_scope(core_all, month)
    previous = previous_month_row(core_all, month)
    products = month_scope(sheets.get("商品经营", pd.DataFrame()), month)
    traffic = month_scope(sheets.get("店铺流量来源", pd.DataFrame()), month)
    product_traffic = month_scope(sheets.get("商品流量来源", pd.DataFrame()), month)
    monthly_ad = month_scope(ad_metrics.get("monthly_summary", pd.DataFrame()), month)
    previous_monthly_ad = previous_month_row(ad_metrics.get("monthly_summary", pd.DataFrame()), month)
    plan_ad = month_scope(ad_metrics.get("plan_detail", pd.DataFrame()), month)
    if not plan_ad.empty and "花费" in plan_ad.columns:
        plan_ad = plan_ad[pd.to_numeric(plan_ad["花费"], errors="coerce").fillna(0) > 0]
        plan_ad = plan_ad.sort_values("花费", ascending=False)
    product_ad = month_scope(ad_metrics.get("product_ad", pd.DataFrame()), month)
    monthly_ad_record = complete_monthly_ad_record(first_record(monthly_ad), first_record(core), plan_ad)
    previous_plan_ad = previous_month_rows(ad_metrics.get("plan_detail", pd.DataFrame()), month)
    previous_monthly_ad_record = complete_monthly_ad_record(first_record(previous_monthly_ad), first_record(previous), previous_plan_ad)

    top_products = select_major_products(products, product_ad, product_traffic)

    return {
        "month": month,
        "core": first_record(core),
        "previous_core": first_record(previous),
        "top_products": records(top_products),
        "top_traffic": records(traffic.sort_values("支付金额", ascending=False).head(8)) if not traffic.empty and "支付金额" in traffic.columns else [],
        "product_traffic": records(product_traffic.sort_values("支付金额", ascending=False).head(12)) if not product_traffic.empty and "支付金额" in product_traffic.columns else [],
        "monthly_ad": monthly_ad_record,
        "previous_monthly_ad": previous_monthly_ad_record,
        "plan_ad": records(plan_ad.head(20)),
        "product_ad": records(product_ad),
        "validation": validation,
    }


def complete_monthly_ad_record(ad_record: dict[str, Any], core_record: dict[str, Any], plan_rows: pd.DataFrame | None = None) -> dict[str, Any]:
    record = dict(ad_record)
    plan_rows = plan_rows if plan_rows is not None else pd.DataFrame()
    if not plan_rows.empty:
        if safe_number(record.get("总推广花费")) <= 0 and "花费" in plan_rows.columns:
            record["总推广花费"] = numeric_sum(plan_rows, "花费")
            record["广告花费来源"] = "推广计划明细"
        if safe_number(record.get("点击量")) <= 0 and "点击量" in plan_rows.columns:
            record["点击量"] = numeric_sum(plan_rows, "点击量")
        if safe_number(record.get("广告引导成交金额")) <= 0:
            if "ROI成交金额_按投入产出比" in plan_rows.columns:
                record["广告引导成交金额"] = numeric_sum(plan_rows, "ROI成交金额_按投入产出比")
            elif {"花费", "ROI"}.issubset(plan_rows.columns):
                spend = pd.to_numeric(plan_rows["花费"], errors="coerce").fillna(0)
                roi = pd.to_numeric(plan_rows["ROI"], errors="coerce").fillna(0)
                record["广告引导成交金额"] = float((spend * roi).sum())
    core_spend = safe_number(core_record.get("推广花费合计"))
    if core_spend > 0 and safe_number(record.get("总推广花费")) <= 0:
        record["总推广花费"] = core_spend
        record["广告花费来源"] = "店铺经营核心月报"
    if safe_number(record.get("总ROI")) <= 0:
        record["总ROI"] = safe_div(safe_number(record.get("广告引导成交金额")), safe_number(record.get("总推广花费")))
    if core_record.get("月份") and not record.get("月份"):
        record["月份"] = core_record.get("月份")
    return record


def select_major_products(products: pd.DataFrame, product_ad: pd.DataFrame, product_traffic: pd.DataFrame | None = None) -> pd.DataFrame:
    product_traffic = product_traffic if product_traffic is not None else pd.DataFrame()
    if not product_ad.empty and "商品名称" in product_ad.columns:
        source = product_ad.copy()
    else:
        source = products.copy()
    if source.empty:
        source = product_rows_from_ad_or_traffic(product_ad, product_traffic)
    if source.empty:
        return source
    if "商品ID" in source.columns:
        source["商品ID_标准"] = source["商品ID"].map(normalize_id)
    else:
        source["商品ID_标准"] = ""
    ad = product_ad.copy()
    if not ad.empty and {"商品ID", "广告花费"}.issubset(ad.columns):
        ad["商品ID_标准"] = ad["商品ID"].map(normalize_id)
        aggregations = {"广告花费": "sum"}
        if "广告ROI" in ad.columns:
            aggregations["广告ROI"] = "max"
        ad = ad.groupby("商品ID_标准", as_index=False).agg(aggregations)
        if source["商品ID_标准"].astype(str).ne("").any():
            source = source.merge(ad, on="商品ID_标准", how="left", suffixes=("", "_广告"))
        elif "商品名称" in source.columns and "商品名称" in ad.columns:
            source = source.merge(ad.drop(columns=["商品ID_标准"], errors="ignore"), on="商品名称", how="left", suffixes=("", "_广告"))
    if "广告花费" not in source.columns:
        source["广告花费"] = 0.0
    if "广告ROI" not in source.columns:
        source["广告ROI"] = 0.0
    scored_fields = ["广告花费", "广告ROI"]
    for field in scored_fields:
        if field in source.columns:
            source[f"{field}_原始缺失"] = source[field].isna()
    source["广告花费"] = pd.to_numeric(source.get("广告花费", 0), errors="coerce").fillna(0)
    source["广告ROI"] = pd.to_numeric(source.get("广告ROI", 0), errors="coerce").fillna(0)
    source["商品价值分"] = (
        percentile_score(source["广告花费"]) * 0.55
        + percentile_score(source["广告ROI"]) * 0.45
    )
    source = source.sort_values(["商品价值分", "广告花费", "广告ROI"], ascending=False)

    selected = source.head(4).copy()
    for field in scored_fields:
        missing_column = f"{field}_原始缺失"
        if missing_column in selected.columns:
            selected.loc[selected[missing_column].fillna(False), field] = None
    return selected.drop(columns=["商品ID_标准", "商品价值分", *[f"{field}_原始缺失" for field in scored_fields]], errors="ignore")


def product_rows_from_ad_or_traffic(product_ad: pd.DataFrame, product_traffic: pd.DataFrame) -> pd.DataFrame:
    if not product_ad.empty and "商品名称" in product_ad.columns:
        rows = product_ad.copy()
        for column in ["广告花费", "广告ROI"]:
            if column not in rows.columns:
                rows[column] = None
        return rows
    if not product_traffic.empty and "商品名称" in product_traffic.columns:
        rows = product_traffic.copy()
        for column in ["商品ID", "广告花费", "广告ROI"]:
            if column not in rows.columns:
                rows[column] = None
        return rows
    return pd.DataFrame()


def percentile_score(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce").fillna(0)
    if values.max() <= 0:
        return pd.Series(0.0, index=series.index)
    return values.rank(pct=True, method="max")


def build_ai_packet(context: dict[str, Any]) -> dict[str, Any]:
    return {
        "role": "你是一名资深 Python 数据工程师、BI分析师和电商数据产品经理。",
        "rules": [
            "只能基于提供的JSON数据分析，不得编造数字。",
            "数值字段由程序计算填充，AI只填写文字字段。",
            "输出必须是JSON，不要Markdown。",
            "不要生成PPT。",
            "3.2 本月主要操作记录跳过。",
        ],
        "expected_json_schema": {
            "product_actions": [{"商品名称": "string", "操作建议": "string"}],
            "goal_paths": {"月销售额或本周期销售额": "string", "广告ROI": "string", "访客数": "string", "转化率": "string"},
            "budget_usage": [{"推广渠道": "关键词推广/全站推广/人群推广", "用途说明": "string"}],
            "weekly_plan": [{"时间节点": "string", "具体内容": "string", "预期效果": "string"}],
            "summary": {"本月亮点": ["string"], "存在问题": ["string"], "下月重点关注": ["string"]},
        },
        "data": context,
    }


def fill_template(template_path: Path, output_path: Path, context: dict[str, Any], ai_text: dict[str, Any]) -> None:
    workbook = load_workbook(template_path)
    sheet = workbook.active
    ensure_store_revenue_section(sheet)
    delete_operation_record_block(sheet)
    month = context["month"]
    period_type = context.get("period_type", "month")
    core = context["core"]
    prev = context["previous_core"]
    monthly_ad = context["monthly_ad"]
    previous_monthly_ad = context["previous_monthly_ad"]
    plan_rows = context["plan_ad"]
    product_rows = context["top_products"]
    product_ad = {normalize_id(row.get("商品ID", "")): row for row in context["product_ad"]}
    should_renumber_sections = False

    sheet["A1"] = report_title(core.get("店铺名称", ""), month, period_type, core)
    metric_header = find_row(sheet, "指标", default=3)
    apply_period_section_headers(sheet, metric_header, period_type)
    write_metric(sheet, metric_header + 1, core, prev, "访客数")
    write_metric(sheet, metric_header + 2, core, prev, "支付金额")
    write_metric(sheet, metric_header + 3, core, prev, "支付转化率_计算", percent=True)
    write_ad_metric(sheet, metric_header + 4, monthly_ad, previous_monthly_ad, "总推广花费")
    write_ad_metric(sheet, metric_header + 5, monthly_ad, previous_monthly_ad, "广告引导成交金额")
    write_ad_metric(sheet, metric_header + 6, monthly_ad, previous_monthly_ad, "总ROI")
    cpc = safe_div(optional_number(monthly_ad, "总推广花费") or 0, optional_number(monthly_ad, "点击量") or 0)
    previous_cpc = safe_div(
        optional_number(previous_monthly_ad, "总推广花费") or 0,
        optional_number(previous_monthly_ad, "点击量") or 0,
    )
    cpc_row = metric_header + 7
    click_row = metric_header + 8
    set_cell(sheet, cpc_row, 2, cpc)
    set_cell(sheet, cpc_row, 3, previous_cpc)
    set_cell(sheet, cpc_row, 4, percent_text(safe_change(cpc, previous_cpc)))
    set_cell(sheet, cpc_row, 5, metric_explanation(sheet.cell(cpc_row, 1).value, cpc, previous_cpc))
    make_cell_not_bold(sheet, cpc_row, 5)
    current_clicks = optional_number(monthly_ad, "点击量")
    previous_clicks = optional_number(previous_monthly_ad, "点击量")
    set_cell(sheet, click_row, 2, current_clicks)
    set_cell(sheet, click_row, 3, previous_clicks)
    set_cell(sheet, click_row, 4, percent_text(safe_change(current_clicks, previous_clicks)))
    set_cell(sheet, click_row, 5, metric_explanation(
        sheet.cell(click_row, 1).value,
        current_clicks,
        previous_clicks,
    ))
    make_cell_not_bold(sheet, click_row, 5)

    if product_rows:
        product_start = find_row(sheet, "商品名称", default=13) + 1
        product_end = first_existing_row(sheet, ["三、", "四、"], product_start, default=product_start + 6)
        product_capacity = max(0, product_end - product_start)
        product_count = min(4, len(product_rows), product_capacity)
        clear_rows(sheet, product_start, product_end - 1, 1, 8)
        product_actions = {item.get("商品名称"): item.get("操作建议") for item in ai_text.get("product_actions", [])}
        for offset, product in enumerate(product_rows[:product_count], start=product_start):
            product_id = normalize_id(product.get("商品ID", ""))
            ad = product_ad.get(product_id, {})
            set_cell(sheet, offset, 1, product.get("商品名称"))
            set_cell(sheet, offset, 2, ad.get("广告花费", product.get("广告花费")))
            set_cell(sheet, offset, 3, ad.get("广告ROI", product.get("广告ROI")))
            set_cell(sheet, offset, 4, product_actions.get(product.get("商品名称"), ""))
    else:
        product_start = find_row(sheet, "商品名称", default=13) + 1
        product_end = first_existing_row(sheet, ["三、", "四、"], product_start, default=product_start + 6)
        clear_rows(sheet, product_start, product_end - 1, 1, 8)

    plan_start = find_row(sheet, "推广计划名称", default=21) + 1
    plan_section_end = first_existing_row(sheet, ["四、", "五、"], plan_start, default=plan_start + 5)
    total_row = find_total_row(sheet, plan_start, plan_section_end)
    detail_end = max(plan_start, total_row - 1)
    plan_capacity = max(0, detail_end - plan_start + 1)
    clear_rows(sheet, plan_start, detail_end, 1, 8)
    visible_plans = plan_rows[:plan_capacity]
    total_cost = 0.0
    total_impressions = 0.0
    total_clicks = 0.0
    weighted_roi_numerator = 0.0
    for offset, plan in enumerate(visible_plans, start=plan_start):
        cost = safe_number(plan.get("花费"))
        clicks = safe_number(plan.get("点击量"))
        impressions = safe_number(plan.get("展现量"))
        roi = safe_number(plan.get("ROI"))
        total_cost += cost
        total_impressions += impressions
        total_clicks += clicks
        weighted_roi_numerator += cost * roi
        set_cell(sheet, offset, 1, plan.get("推广计划名称"))
        set_cell(sheet, offset, 2, plan.get("计划类型"))
        set_cell(sheet, offset, 3, cost)
        set_cell(sheet, offset, 4, impressions)
        set_cell(sheet, offset, 5, whole_number(clicks))
        set_cell(sheet, offset, 6, percent_text(clicks / impressions if impressions else None))
        set_cell(sheet, offset, 7, cost / clicks if clicks else None)
        set_cell(sheet, offset, 8, roi)
        copy_font(sheet, offset, 7, offset, 8)

    set_cell(sheet, total_row, 1, "合计")
    set_cell(sheet, total_row, 3, total_cost)
    set_cell(sheet, total_row, 4, total_impressions)
    set_cell(sheet, total_row, 5, whole_number(total_clicks))
    set_cell(sheet, total_row, 6, percent_text(total_clicks / total_impressions if total_impressions else None))
    set_cell(sheet, total_row, 7, total_cost / total_clicks if total_clicks else None)
    set_cell(sheet, total_row, 8, weighted_roi_numerator / total_cost if total_cost else None)
    copy_font(sheet, total_row, 7, total_row, 8)
    fill_planning(sheet, context, ai_text)
    fill_summary(sheet, ai_text)
    if should_renumber_sections:
        renumber_sections_after_deleted_product(sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_metric(sheet, row: int, core: dict[str, Any], previous: dict[str, Any], field: str, percent: bool = False) -> None:
    current_value = safe_number(core.get(field))
    previous_value = safe_number(previous.get(field))
    set_cell(sheet, row, 2, percent_text(current_value) if percent else current_value)
    set_cell(sheet, row, 3, percent_text(previous_value) if percent else previous_value)
    set_cell(sheet, row, 4, percent_text((current_value - previous_value) / previous_value) if previous_value else "")
    set_cell(sheet, row, 5, metric_explanation(sheet.cell(row, 1).value, current_value, previous_value))
    make_cell_not_bold(sheet, row, 5)


def write_ad_metric(sheet, row: int, monthly_ad: dict[str, Any], previous_ad: dict[str, Any], field: str) -> None:
    current_value = optional_number(monthly_ad, field)
    previous_value = optional_number(previous_ad, field)
    set_cell(sheet, row, 2, current_value)
    set_cell(sheet, row, 3, previous_value)
    set_cell(sheet, row, 4, percent_text(safe_change(current_value, previous_value)))
    set_cell(sheet, row, 5, metric_explanation(sheet.cell(row, 1).value, current_value, previous_value))
    make_cell_not_bold(sheet, row, 5)


def fill_planning(sheet, context: dict[str, Any], ai_text: dict[str, Any]) -> None:
    core = context["core"]
    period_type = context.get("period_type", "month")
    monthly_ad = context["monthly_ad"]
    previous_ad = context["previous_monthly_ad"]
    plan_rows = context["plan_ad"]
    sales_current = safe_number(core.get("支付金额"))
    visitors_current = safe_number(core.get("访客数"))
    conversion_current = safe_number(core.get("支付转化率_计算"))
    roi_current = safe_number(monthly_ad.get("总ROI"))
    roi_previous = safe_number(previous_ad.get("总ROI"))

    sales_label = period_sales_label(period_type)
    next_label = period_next_label(period_type)
    amount_label = period_amount_label(period_type)
    targets = {
        sales_label: round(sales_current * 1.12, 2),
        "广告ROI": round(max(roi_current, roi_previous) * 1.05, 2),
        "访客数": round(visitors_current * 1.08),
        "转化率": round(max(conversion_current * 1.08, conversion_current + 0.001), 4),
    }
    currents = {
        sales_label: sales_current,
        "广告ROI": roi_current,
        "访客数": visitors_current,
        "转化率": conversion_current,
    }
    goal_paths = ai_text.get("goal_paths", {})
    goal_title_row = find_section_row(sheet, "4.1")
    if goal_title_row:
        set_cell(sheet, goal_title_row, 1, f"4.1 {next_label}核心目标")
    goal_header = find_row_after(sheet, "指标", goal_title_row or find_row(sheet, "4.1 下月核心目标", default=30), default=31)
    goal_start = goal_header + 1
    clear_rows(sheet, goal_start, goal_start + 3, 2, 4)
    for row, key in [
        (goal_start, sales_label),
        (goal_start + 1, "广告ROI"),
        (goal_start + 2, "访客数"),
        (goal_start + 3, "转化率"),
    ]:
        if key == sales_label:
            set_cell(sheet, row, 1, f"{sales_label}（元）")
        if key == "转化率":
            set_cell(sheet, row, 2, percent_text(targets[key]))
            set_cell(sheet, row, 3, percent_text(targets[key] - currents[key]))
        else:
            set_cell(sheet, row, 2, targets[key])
            set_cell(sheet, row, 3, targets[key] - currents[key])
        set_cell(sheet, row, 4, goal_paths.get(key, goal_paths.get("月销售额", "") if key == sales_label else ""))

    budget_plan = build_budget_plan(ai_text.get("budget_usage", []), safe_number(monthly_ad.get("总推广花费")))
    budget_title_row = find_section_row(sheet, "4.2") or find_row(sheet, "4.2 下月预算分配计划", default=37)
    if "合计" in budget_plan:
        set_cell(sheet, budget_title_row, 1, f"4.2 {next_label}预算分配计划（{amount_label}{budget_plan['合计'][0]:,.0f}元）")
    else:
        set_cell(sheet, budget_title_row, 1, f"4.2 {next_label}预算分配计划")
    budget_items = [item for item in budget_plan.get("_items", []) if item.get("推广渠道")]
    budget_start = find_row(sheet, "推广渠道", default=38) + 1
    existing_budget_end = next_section_row(sheet, budget_start)
    budget_total_row = find_total_row(sheet, budget_start, existing_budget_end)
    budget_capacity = max(0, budget_total_row - budget_start)
    clear_rows(sheet, budget_start, existing_budget_end - 1, 1, 4)
    visible_budget_items = budget_items[:budget_capacity]
    budget_channels = [item["推广渠道"] for item in visible_budget_items]
    for offset, channel in enumerate(budget_channels):
        row = budget_start + offset
        set_cell(sheet, row, 1, channel)
        if channel in budget_plan:
            amount, ratio = budget_plan[channel]
            set_cell(sheet, row, 2, amount)
            set_cell(sheet, row, 3, ratio_text(ratio))
        else:
            clear_cell(sheet, row, 2)
            clear_cell(sheet, row, 3)
        item = next((entry for entry in visible_budget_items if entry.get("推广渠道") == channel), {})
        set_cell(sheet, row, 4, item.get("用途说明"))
    set_cell(sheet, budget_total_row, 1, "合计")
    if visible_budget_items and "合计" in budget_plan:
        set_cell(sheet, budget_total_row, 2, budget_plan["合计"][0])
        set_cell(sheet, budget_total_row, 3, ratio_text(budget_plan["合计"][1]))
        set_cell(sheet, budget_total_row, 4, "预算按本期投放表现汇总。")

    weekly_items = normalize_weekly_items(ai_text.get("weekly_plan", []), period_type)
    weekly_title_row = find_section_row(sheet, "4.3")
    if weekly_title_row:
        set_cell(sheet, weekly_title_row, 1, planning_section_title(period_type))
    weekly_header = find_row(sheet, "时间节点", default=47)
    content_col = 3
    effect_col = 6
    set_cell(sheet, weekly_header, content_col, "具体内容")
    set_cell(sheet, weekly_header, effect_col, "预期效果")
    weekly_start = weekly_header + 1
    weekly_section_end = next_section_row(sheet, weekly_start)
    weekly_capacity = max(0, weekly_section_end - weekly_start)
    planned_weeks = min(planning_week_count(period_type), weekly_capacity)
    for row in range(weekly_start, weekly_section_end):
        for col in range(2, 9):
            clear_cell(sheet, row, col)
    for row in range(weekly_start, weekly_start + planned_weeks):
        week = sheet.cell(row, 1).value
        item = weekly_items[row - weekly_start] if row - weekly_start < len(weekly_items) else {}
        set_cell(sheet, row, 1, item.get("时间节点", week))
        set_cell(sheet, row, 2, item.get("操作事项"))
        set_cell(sheet, row, content_col, item.get("具体内容"))
        set_cell(sheet, row, effect_col, item.get("预期效果"))


def fill_summary(sheet, ai_text: dict[str, Any]) -> None:
    summary = ai_text.get("summary", {})
    highlight_row = find_summary_label_row(sheet, "亮点")
    problem_row = find_summary_label_row(sheet, "存在问题")
    focus_row = find_summary_label_row(sheet, "重点关注")
    if highlight_row:
        write_lines(sheet, highlight_row + 1, summary.get("本月亮点", []), max_lines=max(0, (problem_row or highlight_row + 6) - highlight_row - 1))
    if problem_row:
        write_lines(sheet, problem_row + 1, summary.get("存在问题", []), max_lines=max(0, (focus_row or problem_row + 6) - problem_row - 1))
    if focus_row:
        write_lines(sheet, focus_row + 1, summary.get("下月重点关注", []), max_lines=5)


def write_lines(sheet, start_row: int, lines: list[str], max_lines: int) -> None:
    for index in range(max_lines):
        set_cell(sheet, start_row + index, 1, lines[index] if index < len(lines) else "")


def ensure_summary_slots(sheet) -> None:
    return


def metric_explanation(label: Any, current: float | None, previous: float | None) -> str:
    text = str(label or "")
    if current is None:
        if "广告" in text or "点击" in text or "ROI" in text:
            return "缺少本月广告明细数据，暂不写0；请补充营销场景/计划/商品广告数据后判断。"
        return "缺少本月数据，暂不判断环比。"
    if not previous:
        return "缺少上月基数，暂不判断环比，只看本月绝对表现。"
    change = current / previous - 1
    signed = f"{change:+.1%}"
    abs_change = f"{abs(change):.1%}"
    if "访客" in text:
        return f"访客环比{signed}，流量入口明显{'放大' if change > 0 else '收缩'}；下一步要看新增流量是否能被转化承接。"
    if "成交金额" in text and "广告带来" not in text:
        return f"成交额较上期{'增加' if change > 0 else '减少'}{abs_change}，变化幅度{'低于' if change < 1 else '高于'}访客增速，说明成交承接仍需单独拆看。"
    if "转化率" in text:
        return f"转化率从{previous:.1%}到{current:.1%}，环比{signed}；若流量上涨同时转化走低，重点排查流量精准度和详情页承接。"
    if "广告总花费" in text:
        return f"广告投入较上期{signed}，预算力度{'加大' if change > 0 else '回收'}；判断是否健康，要同步看广告成交和ROI。"
    if "广告带来" in text:
        return f"广告引导成交环比{signed}，{'跑赢花费增速，投放拉动有效' if change > 0 else '低于预期，需要回看计划成交质量'}。"
    if "ROI" in text:
        return f"ROI本期{current:.1f}、上期{previous:.1f}，{'效率小幅改善' if change > 0 else '效率回落'}，预算优先向稳定高ROI计划倾斜。"
    if "平均点击成本" in text:
        return f"CPC较上期{signed}，单次点击成本{'抬升' if change > 0 else '下降'}；若点击同步增长，说明流量采购效率更好。"
    if "点击量" in text:
        return f"点击量扩大{signed}，广告触达规模变化明显；后续重点看点击增长是否同步沉淀为成交。"
    return f"本期较上期{signed}，建议结合上下游指标继续判断。"


def clear_rows(sheet, start: int, end: int, start_col: int, end_col: int) -> None:
    for row in range(start, end + 1):
        for col in range(start_col, end_col + 1):
            clear_cell(sheet, row, col)


def writable_cell(sheet, row: int, col: int):
    cell = sheet.cell(row, col)
    if not isinstance(cell, MergedCell):
        return cell
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return sheet.cell(merged.min_row, merged.min_col)
    return cell


def set_cell(sheet, row: int, col: int, value: Any) -> None:
    cell = writable_cell(sheet, row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = display_value(value)


def make_cell_not_bold(sheet, row: int, col: int) -> None:
    cell = writable_cell(sheet, row, col)
    if isinstance(cell, MergedCell):
        return
    font = copy.copy(cell.font)
    font.bold = False
    cell.font = font


def copy_font(sheet, source_row: int, source_col: int, target_row: int, target_col: int) -> None:
    source = writable_cell(sheet, source_row, source_col)
    target = writable_cell(sheet, target_row, target_col)
    if isinstance(source, MergedCell) or isinstance(target, MergedCell):
        return
    target.font = copy.copy(source.font)


def clear_cell(sheet, row: int, col: int) -> None:
    cell = sheet.cell(row, col)
    if isinstance(cell, MergedCell):
        return
    cell.value = None


def delete_empty_rows_after(sheet, row: int, max_col: int) -> None:
    return


def insert_blank_rows_after(sheet, row: int, amount: int) -> None:
    return


def ensure_two_blank_rows_between_major_sections(sheet) -> None:
    return


def count_blank_rows_before(sheet, row: int) -> int:
    count = 0
    cursor = row - 1
    while cursor >= 1:
        if any(sheet.cell(cursor, col).value not in (None, "") for col in range(1, sheet.max_column + 1)):
            break
        count += 1
        cursor -= 1
    return count


def find_row(sheet, needle: str, default: int) -> int:
    for row in range(1, sheet.max_row + 1):
        value = str(sheet.cell(row, 1).value or "").strip()
        if needle in value:
            return row
    return default


def find_row_after(sheet, needle: str, after_row: int, default: int) -> int:
    for row in range(after_row + 1, sheet.max_row + 1):
        value = str(sheet.cell(row, 1).value or "").strip()
        if needle in value:
            return row
    return default


def find_total_row(sheet, start: int, end: int) -> int:
    for row in range(start, max(start, end)):
        value = str(sheet.cell(row, 1).value or "").strip()
        if value == "合计":
            return row
    return max(start, end - 1)


def find_summary_label_row(sheet, needle: str) -> int:
    summary_start = find_row_by_any_cell(sheet, "运营总结与建议")
    start = summary_start or 1
    for row in range(start, sheet.max_row + 1):
        value = str(sheet.cell(row, 1).value or "").strip()
        if needle in value and value.endswith("："):
            return row
    return 0


def delete_operation_record_block(sheet) -> None:
    return


def ensure_store_revenue_section(sheet) -> None:
    return


def apply_period_section_headers(sheet, metric_header: int, period_type: str) -> None:
    section_titles = {
        "week": "一、近7天店铺收益",
        "half-month": "一、近15天店铺收益",
        "month": "一、店铺收益",
    }
    current_headers = {
        "week": "近7天数值",
        "half-month": "近15天数值",
        "month": "本月数值",
    }
    previous_headers = {
        "week": "前7天数值",
        "half-month": "前15天数值",
        "month": "上月数值",
    }
    set_cell(sheet, 2, 1, section_titles.get(period_type, "一、店铺收益"))
    set_cell(sheet, metric_header, 2, current_headers.get(period_type, "本期数值"))
    set_cell(sheet, metric_header, 3, previous_headers.get(period_type, "上期数值"))


def delete_section(sheet, start_title: str, next_title: str) -> None:
    return


def ensure_product_rows(sheet, product_start: int, product_count: int) -> None:
    return


def copy_row_style(sheet, source_row: int, target_row: int) -> None:
    return


def renumber_sections_after_deleted_product(sheet) -> None:
    replacements = [
        ("三、", "二、"),
        ("四、", "三、"),
        ("五、", "四、"),
        ("3.", "2."),
        ("4.", "3."),
    ]
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if not isinstance(value, str):
            continue
        stripped = value.strip()
        for old, new in replacements:
            if stripped.startswith(old):
                set_cell(sheet, row, 1, new + stripped[len(old):])
                break


def default_budget_usage(channel: Any) -> str:
    mapping = {
        "关键词推广": "用于承接核心词、长尾词和流量智选流量，优先保留高ROI词包；连续低效词降价或暂停。",
        "全站推广": "用于承接成交款放量，每日按ROI、CPC和库存情况微调。",
        "人群推广": "用于测试高意向人群和复购人群，低效人群先降溢价，单次不超过30%。",
    }
    return mapping.get(str(channel), "")


def ratio_text(value: Any) -> str:
    return percent_text(value)


def percent_text(value: Any) -> str:
    try:
        ratio = float(value)
    except (TypeError, ValueError):
        return ""
    if pd.isna(ratio):
        return ""
    return f"{trim_number(ratio * 100, 2)}%"


def display_value(value: Any) -> Any:
    if isinstance(value, bool) or value is None:
        return value
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        rounded = round(float(value), 2)
        return int(rounded) if rounded.is_integer() else rounded
    return value


def whole_number(value: Any) -> int | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if pd.isna(number):
        return None
    return int(round(number))


def trim_number(value: float, max_digits: int = 2) -> str:
    text = f"{value:.{max_digits}f}"
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def planning_section_title(period_type: str) -> str:
    return {
        "week": "4.3 下周重点操作规划",
        "half-month": "4.3 下半月重点操作规划",
        "month": "4.3 下月重点操作规划",
    }.get(period_type, "4.3 下月重点操作规划")


def planning_week_count(period_type: str) -> int:
    return {
        "week": 1,
        "half-month": 2,
        "month": 4,
    }.get(period_type, 4)


def period_name(period_type: str) -> str:
    return {"week": "周度", "half-month": "半月", "month": "月度"}.get(period_type, "月度")


def period_next_label(period_type: str) -> str:
    return {"week": "下周", "half-month": "下半月", "month": "下月"}.get(period_type, "下月")


def period_amount_label(period_type: str) -> str:
    return {"week": "周约", "half-month": "半月约", "month": "月约"}.get(period_type, "月约")


def period_sales_label(period_type: str) -> str:
    return {"week": "周销售额", "half-month": "半月销售额", "month": "月销售额"}.get(period_type, "月销售额")


def normalize_period_wording(sheet, period_type: str) -> None:
    return


def report_title(store_name: Any, period_value: str, period_type: str, core_record: dict[str, Any] | None = None) -> str:
    suffix = {"week": "周报", "half-month": "半月报", "month": "月报"}.get(period_type, "月报")
    period_text, _is_partial = display_period_for_title(period_value, period_type, core_record)
    return f"{store_name}{period_text}{suffix}"


def display_period_for_title(period_value: str, period_type: str, core_record: dict[str, Any] | None = None) -> tuple[str, bool]:
    text = str(period_value)
    if "~" in text:
        start_text, end_text = text.split("~", 1)
        start = pd.to_datetime(start_text, errors="coerce")
        end = pd.to_datetime(end_text, errors="coerce")
        if not pd.isna(start) and not pd.isna(end):
            if start.year == end.year:
                return f"{start.year}年{start.month:02d}月{start.day:02d}日至{end.month:02d}月{end.day:02d}日", False
            return f"{start.year}年{start.month:02d}月{start.day:02d}日至{end.year}年{end.month:02d}月{end.day:02d}日", False
    if period_type == "month" and len(text) == 7 and text[4] == "-":
        return f"{text[:4]}年{text[5:7]}月", False
    if period_type == "half-month" and len(text) >= 10 and text[4] == "-":
        return f"{text[:4]}年{text[5:7]}月{text[7:]}", False
    if period_type == "week":
        return text.replace("~", "至"), False
    return text, False


def display_range_period(core_record: dict[str, Any] | None) -> str:
    if not core_record:
        return ""
    start = pd.to_datetime(core_record.get("起始日期"), errors="coerce")
    end = pd.to_datetime(core_record.get("结束日期"), errors="coerce")
    if pd.isna(start) or pd.isna(end):
        return ""
    month_start = start.replace(day=1)
    month_end = month_start + pd.offsets.MonthEnd(0)
    if start.normalize() == month_start.normalize() and end.normalize() == month_end.normalize():
        return ""
    if start.year == end.year:
        return f"{start.year}年{start.month:02d}月{start.day:02d}日至{end.month:02d}月{end.day:02d}日"
    return f"{start.year}年{start.month:02d}月{start.day:02d}日至{end.year}年{end.month:02d}月{end.day:02d}日"


def apply_final_number_formats(sheet) -> None:
    return


def apply_core_metric_formats(sheet) -> None:
    return


def apply_product_formats(sheet) -> None:
    return


def apply_plan_formats(sheet) -> None:
    return


def apply_goal_formats(sheet) -> None:
    return


def apply_budget_formats(sheet) -> None:
    return


def repair_budget_total_row(sheet) -> None:
    return


def metric_number_format(label: str) -> str:
    if "转化率" in label or "点击率" in label or "占比" in label or "费比" in label:
        return "0.0%"
    if "ROI" in label or "投入产出比" in label or "平均点击成本" in label:
        return "0.0"
    return "#,##0"


def first_existing_row(sheet, prefixes: list[str], start: int, default: int) -> int:
    for row in range(start, sheet.max_row + 1):
        value = str(sheet.cell(row, 1).value or "").strip()
        if any(value.startswith(prefix) for prefix in prefixes):
            return row
    return default


def apply_final_layout(sheet) -> None:
    return


def header_and_section_rows(sheet) -> set[int]:
    rows: set[int] = {1}
    for row in range(1, sheet.max_row + 1):
        first = str(sheet.cell(row, 1).value or "").strip()
        if first.startswith(("一、", "二、", "三、", "四、", "五、")) or first.startswith(("2.", "3.", "4.")):
            rows.add(row)
        elif first in {"指标", "商品名称", "推广计划名称", "推广渠道", "时间节点"}:
            rows.add(row)
    return rows


def apply_summary_styles(sheet, text_font: str) -> None:
    return


def find_row_by_any_cell(sheet, needle: str) -> int:
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if needle in str(sheet.cell(row, col).value or ""):
                return row
    return 0


def find_section_row(sheet, prefix: str) -> int:
    for row in range(1, sheet.max_row + 1):
        value = str(sheet.cell(row, 1).value or "").strip()
        if value.startswith(prefix):
            return row
    return 0


def next_section_row(sheet, start_row: int) -> int:
    for row in range(start_row + 1, sheet.max_row + 1):
        first = str(sheet.cell(row, 1).value or "").strip()
        if first.startswith(("一、", "二、", "三、", "四、", "五、")) or first.startswith(("2.", "3.", "4.")):
            return row
    return sheet.max_row + 1


def merged_width(sheet, row: int, col: int, widths: dict[str, int]) -> float:
    for merged in sheet.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return sum(widths.get(get_column_letter(index), 14) for index in range(merged.min_col, merged.max_col + 1))
    return widths.get(get_column_letter(col), 14)


def load_ai_json(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {}
    return json.loads(path.read_text(encoding="utf-8"))


def empty_ai_text() -> dict[str, Any]:
    return {
        "product_actions": [],
        "goal_paths": {},
        "budget_usage": [],
        "weekly_plan": [],
        "summary": {"本月亮点": [], "存在问题": [], "下月重点关注": []},
    }


def month_scope(frame: pd.DataFrame, month: str) -> pd.DataFrame:
    if frame.empty or "月份" not in frame.columns:
        return frame.iloc[0:0] if not frame.empty else frame
    return frame[frame["月份"].astype(str).eq(month)].copy()


def previous_month_row(frame: pd.DataFrame, month: str) -> pd.DataFrame:
    if frame.empty or "月份" not in frame.columns:
        return frame.iloc[0:0] if not frame.empty else frame
    scoped = frame[frame["月份"].astype(str).lt(month)].sort_values("月份")
    return scoped.tail(1)


def previous_month_rows(frame: pd.DataFrame, month: str) -> pd.DataFrame:
    if frame.empty or "月份" not in frame.columns:
        return frame.iloc[0:0] if not frame.empty else frame
    values = sorted(str(item) for item in frame["月份"].dropna().astype(str).unique() if str(item) < str(month))
    if not values:
        return frame.iloc[0:0]
    return frame[frame["月份"].astype(str).eq(values[-1])].copy()


def first_record(frame: pd.DataFrame) -> dict[str, Any]:
    if frame.empty:
        return {}
    return clean_record(frame.iloc[0].to_dict())


def records(frame: pd.DataFrame) -> list[dict[str, Any]]:
    if frame.empty:
        return []
    return [clean_record(row) for row in frame.to_dict(orient="records")]


def clean_record(record: dict[str, Any]) -> dict[str, Any]:
    cleaned: dict[str, Any] = {}
    for key, value in record.items():
        if pd.isna(value):
            cleaned[str(key)] = None
        elif isinstance(value, pd.Timestamp):
            cleaned[str(key)] = value.strftime("%Y-%m-%d")
        elif hasattr(value, "item"):
            cleaned[str(key)] = value.item()
        else:
            cleaned[str(key)] = value
    return cleaned


def numeric_sum(frame: pd.DataFrame, column: str) -> float:
    if frame.empty or column not in frame.columns:
        return 0.0
    return float(pd.to_numeric(frame[column], errors="coerce").fillna(0).sum())


def safe_number(value: Any) -> float:
    if value is None or pd.isna(value):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def optional_number(record: dict[str, Any], field: str) -> float | None:
    if field not in record:
        return None
    value = record.get(field)
    if value is None or pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def safe_div(numerator: float, denominator: float) -> float | None:
    return numerator / denominator if denominator else None


def safe_change(current: float | None, previous: float | None) -> float | None:
    if current is None or previous in (None, 0):
        return None
    return (current - previous) / previous


def build_budget_plan(ai_budget_items: list[dict[str, Any]], current_total_spend: float) -> dict[str, Any]:
    items = [item for item in ai_budget_items if isinstance(item, dict) and item.get("推广渠道")]
    if not items:
        return {}
    next_total = round(current_total_spend * 1.05, 2) if current_total_spend > 0 else 0
    ratios = normalize_budget_ratios(items)
    plan: dict[str, Any] = {"_items": items}
    total_amount = 0.0
    total_ratio = 0.0
    for item, ratio in zip(items, ratios, strict=False):
        channel = str(item.get("推广渠道") or "").strip()
        amount = optional_number(item, "预算")
        if amount is None:
            amount = optional_number(item, "预算（元）")
        if amount is None and next_total > 0:
            amount = round(next_total * ratio, 2)
        if amount is not None:
            total_amount += float(amount)
        total_ratio += ratio
        plan[channel] = (amount, ratio)
    plan["合计"] = (round(total_amount, 2) if total_amount else None, round(total_ratio, 4) if total_ratio else None)
    return plan


def normalize_budget_ratios(items: list[dict[str, Any]]) -> list[float]:
    raw: list[float | None] = []
    for item in items:
        ratio = optional_number(item, "占比")
        if ratio is not None and ratio > 1:
            ratio = ratio / 100
        raw.append(ratio)
    if all(value is not None and value >= 0 for value in raw) and sum(value or 0 for value in raw) > 0:
        total = sum(value or 0 for value in raw)
        return [(value or 0) / total for value in raw]
    if not items:
        return []
    equal = 1 / len(items)
    return [equal for _ in items]


def normalize_weekly_items(items: list[dict[str, Any]], period_type: str) -> list[dict[str, Any]]:
    if not isinstance(items, list):
        return []
    count = planning_week_count(period_type)
    normalized: list[dict[str, Any]] = []
    for index, item in enumerate(items[:count], start=1):
        if not isinstance(item, dict):
            continue
        normalized.append(
            {
                "时间节点": item.get("时间节点") or f"第{index}周",
                "操作事项": item.get("操作事项"),
                "具体内容": item.get("具体内容"),
                "预期效果": item.get("预期效果"),
            }
        )
    return normalized


def normalize_id(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip().replace(",", "")
    if text.endswith(".0"):
        text = text[:-2]
    return text


if __name__ == "__main__":
    main()
