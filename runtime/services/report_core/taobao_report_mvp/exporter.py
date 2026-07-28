from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .metrics import BuildResult


SHEET_ORDER = [
    "经营总表",
    "店铺流量来源",
    "商品经营",
    "商品投产比日报",
    "商品流量来源",
    "报表识别日志",
    "数据质量",
]


def write_workbook(result: BuildResult, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name in SHEET_ORDER:
            if sheet_name == "报表识别日志":
                frame = result.audit_log
            elif sheet_name == "数据质量":
                frame = result.quality_log
            else:
                frame = result.sheets.get(sheet_name, pd.DataFrame())
            frame.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        for worksheet in workbook.worksheets:
            style_sheet(worksheet)


def style_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                header = worksheet.cell(row=1, column=cell.column).value or ""
                cell.number_format = number_format_for(str(header))
            cell.alignment = Alignment(vertical="center")

    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        width = min(max(max_length + 2, 10), 42)
        worksheet.column_dimensions[get_column_letter(column)].width = width


def number_format_for(header: str) -> str:
    if "率" in header or "费比" in header or header.endswith("_环比"):
        return "0.00%"
    if "ROI" in header:
        return "0.00"
    if "金额" in header or "花费" in header or "客单价" in header or "UV价值" in header:
        return "#,##0.00"
    return "#,##0"
